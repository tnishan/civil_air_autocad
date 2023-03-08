# import pyautocad
# import streamlit as st
# import pandas as pd
# import sqlite3
# import numpy as np
# import io
# from pyautocad import Autocad, APoint
# # acad= Autocad()
# acad = pyautocad.Autocad(create_if_not_exists=True)
# import xlwings as xw

import comtypes.client;
import os;
import sys;
import pandas as pd
import numpy as np
import math
import openpyxl
import pyautocad
import streamlit as st
import sqlite3
import io
from pyautocad import Autocad, APoint
# acad= Autocad()
acad = pyautocad.Autocad(create_if_not_exists=True)
import xlwings as xw




class AutocadConnection:
    def __init__(self):
     
        
        return
#         self.file_path = file_path
#         self.data = self.load_from_excel()       

    # DRAWS RECTANGLES AND NAMES I.E. FOUNDATIONS INTO THE OPEN AUTOCAD MODEL IN METERS
    def draw_rect(self, width,height,origin_x,origin_y,Name,dimd,txt_ht,col_size,dim_check,name_check):
               
        p1 = APoint(origin_x+width/2, origin_y+height/2)
        p2 = APoint(origin_x+width/2, origin_y-height/2)
        p3 = APoint(origin_x-width/2, origin_y-height/2)
        p4 = APoint(origin_x-width/2, origin_y+height/2)


        line1 = acad.model.AddLine(p1, p2)
        line2 = acad.model.AddLine(p2, p3)
        line3 = acad.model.AddLine(p3, p4)
        line4 = acad.model.AddLine(p4, p1)
        
        line1.layer = 'line_0.25'
        line2.layer = 'line_0.25'
        line3.layer = 'line_0.25'
        line4.layer = 'line_0.25'

        if name_check =='yes':
            t1 = acad.model.AddText(Name, APoint(origin_x+col_size, origin_y+col_size),txt_ht)
            t1.layer = 'Text'
        
        if dim_check =='yes':
            # Add dimensions to the two adjacent lines
            pd1 = APoint(origin_x+width/2+dimd, origin_y+height/2)

            pd2 = APoint(origin_x-width/2, origin_y-height/2-dimd)


            dim1= acad.model.AddDimAligned(p1, p2, pd1)
            dim2 = acad.model.AddDimAligned(p2, p3, pd2)
            dim1.layer = 'Dimension'
            dim2.layer = 'Dimension'

    # DRAWS RECTANGLES AND NAMES I.E. FOUNDATIONS INTO THE OPEN AUTOCAD MODEL IN METERS
    def draw_beams(self, width,origin_x,origin_y,end_x,end_y,Name,txt_ht,orientation):       
        if orientation == 0:
            p1 = APoint(origin_x, origin_y+width/2)
            p2 = APoint(end_x, end_y+width/2)
            p3 = APoint(origin_x, origin_y-width/2)
            p4 = APoint(end_x, end_y-width/2)
            t1 = acad.model.AddText(Name, APoint((origin_x+end_x)/2, origin_y),txt_ht)
            t1.Alignment = 4
            t1.TextAlignmentPoint = APoint((origin_x+end_x)/2, origin_y+width/2+25+txt_ht)
            line1 = acad.model.AddLine(p1, p2)
            line2 = acad.model.AddLine(p3, p4) 

        else:
            p1 = APoint(origin_x+width/2, origin_y)
            p2 = APoint(end_x+width/2, end_y)
            p3 = APoint(origin_x-width/2, origin_y)
            p4 = APoint(end_x-width/2, end_y)
            t1 = acad.model.AddText(Name, APoint(origin_x, (origin_y+end_y)/2),txt_ht)           
            t1.Alignment = 4
            t1.TextAlignmentPoint = APoint(origin_x-width/2-25-txt_ht, (origin_y+end_y)/2)
            t1.rotation = math.pi/2
            line1 = acad.model.AddLine(p1, p2)
            line2 = acad.model.AddLine(p3, p4)          

        line1.layer = 'line_0.25'
        line2.layer = 'line_0.25'
        t1.layer = 'Text'
    
#     ADD BLOCK NAMES TO DRAWING
    def add_block_name(self,Block_origin,offset,txt_ht,block_name,algt):
        t1 = acad.model.AddText(block_name, APoint(Block_origin, 0-offset),txt_ht)
        t1.Alignment = algt
        t1.TextAlignmentPoint = APoint(Block_origin, 0-offset) 
        
    # FUNCTION TO DRAW GRID LINES IN AUTOCAD.
    def draw_grid_lines(self, grid, offset, txt_ht,Block_origin):
    
        # Get the grid names and coordinates
        grid_names_x = list(grid[0].keys())
        grid_names_y = list(grid[1].keys())
        x_coordinates = list(grid[0].values())+Block_origin
        y_coordinates = list(grid[1].values())

        # Get the maximum x and y values
        max_x = max(x_coordinates)
        max_y = max(y_coordinates)

        # Define the origin of the grid
        origin_x = min(x_coordinates) - offset
        origin_y = min(y_coordinates) - offset

        # Loop through each x coordinate
        for i in range(len(x_coordinates)):
            # Create the points for the grid line
            p1 = APoint(x_coordinates[i] , origin_y)
            p2 = APoint(x_coordinates[i] , max_y+offset)

            # Add the line between the two points
            line1 = acad.model.AddLine(p1, p2)
            line1.layer = 'Grid'

            # Add the dimension for the grid line
            if i !=len(x_coordinates)-1:
                pd = APoint(x_coordinates[i+1] , max_y+offset)
                dim1=acad.model.AddDimAligned(APoint(p2.x,p2.y-txt_ht), APoint(pd.x, pd.y - txt_ht),APoint(pd.x, pd.y - txt_ht))
                dim1.layer = 'Dimension'

            # Add the name of the grid line at the end of the line
            acad.model.Addcircle(APoint(p2.x, p2.y + txt_ht), txt_ht)
            t1= acad.model.AddText(grid_names_x[i], APoint(p2.x, p2.y + txt_ht), txt_ht)
            t1.Alignment = 4
            t1.TextAlignmentPoint = APoint(p2.x, p2.y + txt_ht)
            t1.layer = 'Text'

        # Loop through each y coordinate
        for i in range(len(y_coordinates)):
            # Create the points for the grid line
            p1 = APoint(origin_x, y_coordinates[i] )
            p2 = APoint(max_x+offset, y_coordinates[i])

            # Add the line between the two points
            line1 = acad.model.AddLine(p1, p2)
            line1.layer = 'Grid'

        # Add the dimension for the grid line
            if i !=len(y_coordinates)-1:
                pd = APoint(origin_x, y_coordinates[i+1] )
                dim1 = acad.model.AddDimAligned(APoint(p1.x+txt_ht,p1.y), APoint(pd.x+txt_ht, pd.y ),APoint(pd.x+txt_ht, pd.y ))
                dim1.layer = 'Dimension'

            # Add the name of the grid line at the end of the line
            cir1 = acad.model.Addcircle(APoint(p1.x - txt_ht, p1.y), txt_ht)
            cir1.layer = 'Grid'
            
            t2 =acad.model.AddText(grid_names_y[i], APoint(p1.x - txt_ht, p1.y), txt_ht)
            t2.Alignment = 4
            t2.TextAlignmentPoint = APoint(p1.x - txt_ht, p1.y)
            t2.layer = 'Text'
    
    # FUNCTION TO DRAW NUMPY TABLE IN AUTOCAD.
    def draw_table(self,table_from_excel):
        st.subheader("Draw Selected Table from excel to Autocad")
        # if st.button("Import Data From Excel"):
        # table_from_excel = excel_table()
        st.write(table_from_excel)

        col1,col2,col3,col4 = st.columns(4)
        with col1:
            row_ht = st.number_input("Insert Ht. of Rows",value = 50)
        with col2:
            col_ht =  st.number_input("Insert Ht. of Columns",value =200)
        with col3:
            table_header_txt_ht = st.number_input("Insert Ht. of Rows",value = 30)  
        with col4:
            table_content_txt_ht =  st.number_input("Insert Ht. of Columns",value =20)  
        table_header = st.text_input("Insert Table Header",value = table_from_excel[0][0])

        if st.button("Draw Table"):
            # Create the table
            table = acad.model.AddTable(APoint(0,0,0),table_from_excel.shape[0],table_from_excel.shape[1],row_ht,col_ht)

            # iterate over the array and set the text and alignment of each cell in the table
            for i in range(table_from_excel.shape[0]):
                for j in range(table_from_excel.shape[1]):
                    
                    cell_text = str(table_from_excel[i][j]) # convert the element to a string
                    if cell_text != "None":
                        table.SetText(i, j, cell_text) # set the text of the cell
                        table.SetCellAlignment(i, j, 5) # set the alignment of the cell

            table.SetText(0,0,table_header)

            table.SetTextHeight(5,table_content_txt_ht)
            table.SetTextHeight(2,table_header_txt_ht)
            # Update the table with the new text height
            table.Update() 
    
    # FUNCITON TO ADD LAYERS IN AUTOCAD.
    def add_layers(self,lay_list):
        for item in lay_list:
             acad.doc.Layers.Add(item)

    


class ExcelConnection:
    def __init__(self):
        return
#         self.file_path = file_path
#         self.data = self.load_from_excel()       

    def load_from_excel(self,wb_name,ws_name):
        # Load the workbook from the running Excel file
#         wb_name =  wb_name #
       
        wb = openpyxl.load_workbook(wb_name,data_only=True)
        # Select the foundation_design sheet in the workbook
        sheet = wb[ws_name]
        # Get the names of the columns from the third row of the sheet
        all_columns = [sheet.cell(row=4, column=i+1).value for i in range(sheet.max_column) if sheet.cell(row=4, column=i+1).value is not None]
        columns = ['Story', 'Unique Names','Adopted Length', 'Adopted Breadth', 'Type']

        # Create a dictionary to store the data for each column
        data = {col: [] for col in all_columns}

        # Iterate through the rows in the sheet, starting from the fourth row
        for row in sheet.iter_rows(min_row=4, max_col=len(all_columns)):
            # For each column in the row, add the value to the corresponding list in the data dictionary
            for i, cell in enumerate(row):
                data[all_columns[i]].append(cell.value)

        # Create a Pandas DataFrame from the data dictionary
        df = pd.DataFrame(data)
        # Get the list of columns to drop (all column names not in `columns_to_keep`)
        columns_to_drop = [col for col in df.columns if col not in columns]
        # Drop the columns from the DataFrame
        df.drop(columns_to_drop, axis=1, inplace=True)

        # Drop the first row of the DataFrame
        df.drop(index=df.index[:2],inplace=True)
        df.reset_index()
        st.write(df)
        # Convert to mm from m. 
#         print(df)
        df['Adopted Length'] = df['Adopted Length']*1000
        df['Adopted Breadth'] = df['Adopted Breadth']*1000

        df_copy = df.copy()
        # sort values in ascending order of col2 and col3
        df_copy = df_copy.sort_values(by=['Adopted Length', 'Adopted Breadth'])
        df_copy['combined'] = df_copy[['Adopted Length', 'Adopted Breadth']].apply(lambda x: ' '.join(x.astype(str)), axis=1)
        df_unique = df_copy.drop_duplicates(subset='combined')
        df_unique = df_unique.reset_index(drop=True)
        df_unique['assigned_value'] = ["F" + str(i + 1) for i in range(len(df_unique))]
        df_mapping = dict(zip(df_unique['combined'], df_unique['assigned_value']))
        df['Name'] = df_copy['combined'].map(df_mapping)

        return df

    # READ  EXCEL TABLE
    def excel_table(self):
        st.write("CHECK DATA FROM EXCEL")    
        # Get a reference to the active Excel application
        app = xw.apps.active
        # Get a reference to the active worksheet
        sheet = app.books.active.sheets.active
        # Get the selected range of cells
        selection = app.selection

        if selection:
            values = selection.options(ndim=2).value
        else:
            # Get all values in the used range of the worksheet
            values = sheet.used_range.options(ndim=2).value
        
        values = np.array(values)
        # print(values)
        return values
    


# # class EtabsConnection:
#     def __init__(self):
#         return 
#         #         self.acad = Autocad()
    
#     # GETS ESTABLISHES CONNECTION TO ETABS.
#     def connect_to_etabs_2019(self):
#         """
#         Return Values:
#         SapModel (type cOAPI pointer)
#         myETABSObject (type cOAPI pointer)
#         helper (type cOAPI pointer)
#         """
#         #create API helper object
#         helper = comtypes.client.CreateObject('ETABSv1.Helper')
#         helper = helper.QueryInterface(comtypes.gen.ETABSv1.cHelper)
        
#         # #attach to a running instance of ETABS
#         # try:
#         #     #get the active ETABS object
#         #     myETABSObject = helper.GetObject("CSI.ETABS.API.ETABSObject");
#         # except (OSError, comtypes.COMError):
#         #     print("No running instance of the program found or failed to attach.");
#         #     sys.exit(-1);
#         # #create SapModel object
#         myETABSObject = helper.GetObject("CSI.ETABS.API.ETABSObject")
#         SapModel = myETABSObject.SapModel
#         return SapModel,myETABSObject,helper
    
#     # GETS CO-ORDINATES OF ALL THE POINTS FROM ETABS. 
#     def get_coordinates_from_etabs(self,sap_model):
#         # Get all point names
#         point_names = sap_model.PointObj.GetNameList()

#         # Get point coordinates based on unique names
#         point_coords = {}
#         for name in point_names[1]:
#             name = str(name)
#             coord = sap_model.PointObj.GetCoordCartesian(name)
#             point_coords[name] = coord

#         return point_coords

#     # CONVERTS DATA FROM M TO MM
#     def convert_to_mm(self, d):
#         if max(d.values())>1000:
#             print('already in mm')
#             return {k: v * 1 for k, v in d.items()}
#         else:
#             print('m to mm')
#             return {k: v * 1000 for k, v in d.items()}

#     # CONVERTS MM DATA TO FEET ( MAINLY FOR FOUNDATIONS??)   
#     def convert_to_ft(self, d):
#         if max(d.values())>1000:
#             print('mm to ft')
#             return {k: round(v /304.8,3) for k, v in d.items()}
#         else:
#             print('m to ft')
#             return {k: round(v /0.3048,3) for k, v in d.items()}
    
#     # GETS GRID DATA FROM ETABS.
#     def get_etabs_grids(self, sap_model):
#         # GET GRID VALUES
#         get_grids = sap_model.GridSys.GetGridSys_2('G1')

#         grid_x = dict(zip(get_grids[6],get_grids[8] ))
#         grid_y = dict(zip(get_grids[7],get_grids[9] ))
#         grid = [grid_x,grid_y]

#         grid_x_mm = self.convert_to_mm(grid_x)
#         grid_y_mm = self.convert_to_mm(grid_y)
#         grid_mm = [grid_x_mm,grid_y_mm]

#         grid_x_ft = self.convert_to_ft(grid_x)
#         grid_y_ft = self.convert_to_ft(grid_y)
#         grid_ft = [grid_x_ft,grid_y_ft]

#         return grid,grid_mm,grid_ft
    
#     # GETS ALL FRAMES DATA FROM ETABS.
#     def get_allframes(self,sap_model):
#         #RETRIEVES data for all frame objects in the model.
#          all_frames = sap_model.FrameObj.GetAllFrames()
#          return all_frames
    
#     # GETS ALL FRAME PROPERTIES FROM ETABS
#     def GetAllFrameProperties(self,sap_model):
#        #Retrieves select data for all frame properties in the model 
#         all_frame_properties = sap_model.PropFrame.GetAllFrameProperties_2()
#         return all_frame_properties
    
    

# def get_frames():
#     sap_model, etabs_object, helper = EtabsConnection().connect_to_etabs_2019()
#     grid_values, grid_mm, grid_ft = EtabsConnection().get_etabs_grids(sap_model)

#     point_coords = EtabsConnection().get_coordinates_from_etabs(sap_model)
#     # point_coords
#     all_frames = EtabsConnection().get_allframes(sap_model)
#     df_frames = pd.DataFrame(all_frames[1:-1])
#     df_frames = df_frames.transpose()
#     # df_frames = df_frames[['0','1','2','3','4','5','6','7','8','9','10']]
#     df_frames = df_frames.iloc[:, :11]
#     col_names = ['U Name','Section','Story','P1','P2','P1x','P1y','P1z','P2x','P2y','P2z']
#     df_frames.columns = col_names

#     all_frame_properties= EtabsConnection().GetAllFrameProperties(sap_model)
#     df_frame_props = pd.DataFrame(all_frame_properties[1:-1])
#     df_frame_props = df_frame_props.transpose().iloc[:,:4]
#     col_namess = ['Section','ip','Depth','Width']
#     df_frame_props.columns = col_namess
#     df_frame_props = df_frame_props.drop('ip', axis=1)

#     df_frames = pd.merge(df_frames,df_frame_props,on = 'Section')

#     # Define a custom function to extract the first character from the 'Section' column
#     def extract_type(section):
#         return section.split()[0]

#     # Apply the custom function to the 'Section' column
#     df_frames['Type'] = df_frames['Section'].apply(extract_type)

#     # Map the extracted first character to 'Beam', 'Column', or 'Wall'
#     df_frames['Type'] = df_frames['Type'].map({'B': 'Beam', 'C': 'Column', 'W': 'Wall'})


#     # ADD BLOCK ORIGIN
#     df_frames['Story'].unique()

#     # Get unique values from column 'B'
#     unique_values = sorted(df_frames['Story'].unique())

#     # Create a dictionary with unique values as keys and incremental integers as values
#     n = max(df_frames.iloc[:, [5, 6,8,9]].max())*1.5
#     # n = 100000
#     m = 0
#     value_dict = {unique_values[i]: m+i * n for i in range(len(unique_values))}

#     # value_dict
#     df_frames['Block_origin'] = df_frames['Story'].map(value_dict)


#     # ##################ADD ORIENTATION FUNCTION
#     # Calculate the orientation using arctan2
#     df_frames['orientation'] = [math.degrees(math.atan2(P2y - P1y, P2x - P1x)) for P1x, P1y, P2x, P2y in zip(df_frames['P1x'], df_frames['P1y'], df_frames['P2x'], df_frames['P2y'])]

#     # CHANGE TO MM THE LENGTHS FROM M.
#     # df_frames.iloc[:, 5:13] *= 1000
#     # df_frames['Block_origin'] *= 1000

#     st.write(df_frames)
#     return df_frames

# # Define a custom function to be applied to the dataframe to draw beams
# def call_draw_beams(row):   
#     if row['Type'] == 'Beam':
#         AutocadConnection().draw_beams(row['Width'],row['Block_origin'] +row['P1x'],row['P1y'],row['Block_origin'] +row['P2x'],row['P2y'],row['Section'],200,row['orientation'])

# # Define a custom function to be applied to the dataframe to draw columns.
# def call_draw_cols(row):   
#     if row['Type'] == 'Column':
#         AutocadConnection().draw_rect(float( row['Depth']),float( row['Width']), float(row['Block_origin'] + row['P1x']), float(row['P1y']),row['Section'],500,200,500*0.6,'no','no')

# # FUNCTION TO DRAW BEAMS
# def draw_beams():
#     df_frames = get_frames()
#     lay_list = ['line_0.25','Grid','Text','Dimension']
#     AutocadConnection().add_layers(lay_list)
    
#     # df_frames = get_frames()
#     # Use the apply method to apply the custom function to each row
#     df_frames['D'] = df_frames.apply(call_draw_beams, axis=1)

#     for item in df_frames['Story'].unique():
#         print (item)
#         Block_origin = df_frames.loc[df_frames['Story'] == item,'Block_origin'].values[0]
#         AutocadConnection().add_block_name(Block_origin,4000,500,'Beam Layout Plan at  ' + str(item),5)

# FUNCTION TO DRAW COLUMNS
# def draw_columns():
#     connection = EtabsConnection()
#     sap_model, etabs_object, helper = connection.connect_to_etabs_2019()
#     grid_values, grid_mm, grid_ft = connection.get_etabs_grids(sap_model)
    
#     df_frames = get_frames()
#     # Use the apply method to apply the custom function to each row
    
#     df_frames['D'] = df_frames.apply(call_draw_cols, axis=1)


#     # DRAW MULTIPLE GRIDS AS PER REUIREMENT OF COLUNN GRIDS
#     for item in df_frames['Story'].unique():
#         Block_origin = df_frames.loc[df_frames['Story'] == item,'Block_origin'].values[0]
#         AutocadConnection().draw_grid_lines(grid_mm,2000,300,Block_origin)

#FUNCITON SHOWING IN STREAMLIT APP WITH BUTTONS AND INPUTS. 
# def frame_function():
        
#     if st.button("Draw Beams"):
#         draw_beams()

#     if st.button("Draw Columns"):
#         draw_columns()



# def excel_autocad_table():
#     st.subheader("Draw Selected Table from excel to Autocad")
#     # if st.button("Import Data From Excel"):
#     table_from_excel = excel_table()
#     st.write(table_from_excel)

#     col1,col2,col3,col4 = st.columns(4)
#     with col1:
#         row_ht = st.number_input("Insert Ht. of Rows",value = 50)
#     with col2:
#         col_ht =  st.number_input("Insert Ht. of Columns",value =200)
#     with col3:
#         table_header_txt_ht = st.number_input("Insert Ht. of Rows",value = 30)  
#     with col4:
#         table_content_txt_ht =  st.number_input("Insert Ht. of Columns",value =20)  
#     table_header = st.text_input("Insert Table Header",value = table_from_excel[0][0])

#     if st.button("Draw Table"):
#         # Create the table
#         table = acad.model.AddTable(APoint(0,0,0),table_from_excel.shape[0],table_from_excel.shape[1],row_ht,col_ht)

#         # iterate over the array and set the text and alignment of each cell in the table
#         for i in range(table_from_excel.shape[0]):
#             for j in range(table_from_excel.shape[1]):
                
#                 cell_text = str(table_from_excel[i][j]) # convert the element to a string
#                 if cell_text != "None":
#                     table.SetText(i, j, cell_text) # set the text of the cell
#                     table.SetCellAlignment(i, j, 5) # set the alignment of the cell

#         table.SetText(0,0,table_header)

#         table.SetTextHeight(5,table_content_txt_ht)
#         table.SetTextHeight(2,table_header_txt_ht)
#         # Update the table with the new text height
#         table.Update()

# def excel_table():
#     # READ FROM EXCEL TABLE
#     st.write("CHECK DATA FROM EXCEL")    
#     # Get a reference to the active Excel application
#     app = xw.apps.active
#     # Get a reference to the active worksheet
#     sheet = app.books.active.sheets.active
#     # Get the selected range of cells
#     selection = app.selection

#     if selection:
#         values = selection.options(ndim=2).value
#     else:
#         # Get all values in the used range of the worksheet
#         values = sheet.used_range.options(ndim=2).value
    
#     values = np.array(values)
#     # print(values)
#     return values

# ADD LABEL
# def find_xy(label):
#     sap_model, etabs_object, helper = EtabsConnection().connect_to_etabs_2019()
#     point_coords = EtabsConnection().get_coordinates_from_etabs(sap_model)
#     return point_coords[str(label)][0], point_coords[str(label)][1]


# def draw_foundations():
#     if st.button("Draw foundation"):
#         wb_name = 'Darchula_footing_api.xlsx'
#         ws_name = 'Foundation_Design'
#         df = ExcelConnection().load_from_excel(wb_name,ws_name)
#         df = df.dropna()
        
#         # df['Name'].unique().sort_values(by=['Name'])
#         df_names = df['Name'].unique()
#         np.sort(df_names)    
#         df['x'], df['y'] = zip(*df['Unique Names'].apply(find_xy))

#         sap_model, etabs_object, helper = EtabsConnection().connect_to_etabs_2019()
#         grid_values, grid_mm, grid_ft = EtabsConnection().get_etabs_grids(sap_model)

#         # DRAW IN M
#         # DRAW  GRID AS PER THE BLOCK ORIGIN
#         AutocadConnection().draw_grid_lines(grid_mm,2000,300,[])

#         ## FUNCTION TO ADD NAME OF THE BLOCK
#         AutocadConnection().add_block_name(0,4000,500,'FOUNDATION LAYOUT PLAN',5)

#         ## FUNCTION TO CALL THE DRAW RECTANGLE FUNCTION FROM EACH ROW OF PANDA DATAFRAME. 
#         for index, row in df.iterrows():
#             AutocadConnection().draw_rect(float(row['Adopted Length']),float( row['Adopted Breadth']), float(row['x']), float(row['y']),row['Name'],500,200,500*0.6,'yes','yes')



def draw_excel_table():
    AutocadConnection().draw_table(ExcelConnection().excel_table())

def main():  
    st.set_page_config(page_title="CIVIL_AIR")
    st.sidebar.title("Menu")

    menu_options = ["Draw Table", "Draw Frames","Draw Foundation"]
    menu_choice = st.sidebar.selectbox("Select a menu", menu_options)
    if menu_choice == "Draw Table":
       draw_excel_table()

    # if menu_choice == "Draw Frames":
    #    frame_function()
    
    # if menu_choice == "Draw Foundation":
    #    draw_foundations()
       
    
    # Add contact address
    st.sidebar.subheader("Contact:")
    st.sidebar.markdown("Nishan Thapa")
    st.sidebar.markdown("IDEA Consult Pvt. Ltd. - Nepal")
    st.sidebar.markdown("""<img src='https://img.icons8.com/color/48/000000/linkedin.png'
                 style='height:15px;width:15px'/> [Connect on LinkedIn!](https://www.linkedin.com/in/nishan-thapa-b0953362/)""",
            unsafe_allow_html=True)
    
    if st.sidebar.button("HOW TO"):
        # how_to()
        st.write("ITS HOW TO PAGE")

if __name__ == "__main__":
    main()

